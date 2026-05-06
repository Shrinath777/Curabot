"""
CuraBot Golden Test Framework — Diagnostic Accuracy Benchmark Runner

Tests the diagnostic pipeline against 50 pre-validated golden test cases
and generates comprehensive accuracy metrics.

Usage:
  python -m evaluation.run_golden_tests                    # Run all 50 tests
  python -m evaluation.run_golden_tests --limit 10         # Run first 10 tests
  python -m evaluation.run_golden_tests --severity critical # Run only critical cases
  python -m evaluation.run_golden_tests --id GT-001        # Run a single test

Output:
  - Console summary with Top-1, Top-3, and per-severity accuracy
  - Detailed JSON report saved to evaluation/golden_test_results.json
  - Excel summary saved to evaluation/golden_test_report.xlsx
"""

import json
import asyncio
import os
import sys
import logging
import argparse
import time
from datetime import datetime
from typing import Dict, Any, List, Optional

# Add backend to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agents.symptom_normalizer import symptom_normalizer
from agents.hypothesis_generator import hypothesis_generator
from agents.evidence_evaluator import evidence_evaluator
from agents.hypothesis_reviser import hypothesis_reviser

logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)


def load_golden_cases(filter_severity: str = None, filter_id: str = None, limit: int = None) -> List[Dict]:
    """Load and filter golden test cases."""
    cases_path = os.path.join(os.path.dirname(__file__), "golden_test_cases.json")
    with open(cases_path, "r", encoding="utf-8") as f:
        cases = json.load(f)

    if filter_id:
        cases = [c for c in cases if c["id"] == filter_id]
    if filter_severity:
        cases = [c for c in cases if c["severity"] == filter_severity]
    if limit:
        cases = cases[:limit]

    return cases


async def run_single_test(case: Dict) -> Dict[str, Any]:
    """
    Run a single golden test case through the first 4 agents (no LLM for strategist).
    Returns the test result with pass/fail for Top-1 and Top-3 accuracy.
    """
    case_id = case["id"]
    ground_truth = case["ground_truth"]
    message = case["patient_message"]
    expected_top3 = case.get("expected_top3", [])

    start_time = time.time()

    # --- Agent 1: Symptom Normalizer ---
    try:
        normalized = await symptom_normalizer.process(
            message=message,
            conversation_history=[],
            user_context=None,
        )
    except Exception:
        normalized = symptom_normalizer._keyword_fallback(message)

    # --- Agent 2: Hypothesis Generator ---
    try:
        hypothesis_result = await hypothesis_generator.process(
            normalized_symptoms=normalized,
            user_context=None,
            is_new_user=True,
        )
    except Exception:
        hypothesis_result = hypothesis_generator._fallback_hypotheses(normalized)

    hypotheses = hypothesis_result.get("hypotheses", [])

    # --- Agent 3: Evidence Evaluator (rule-based) ---
    all_evidence = []
    for s in normalized.get("primary_symptoms", []):
        all_evidence.append({"type": "symptom", "finding": s.get("name", ""), "detail": ""})
    for s in normalized.get("secondary_symptoms", []):
        all_evidence.append({"type": "symptom", "finding": s.get("name", ""), "detail": ""})

    evidence_result = evidence_evaluator._fallback_evidence(hypotheses, all_evidence)

    # --- Agent 4: Hypothesis Reviser (Bayesian rule-based) ---
    revision_result = hypothesis_reviser._fallback_revision(hypotheses, evidence_result)
    revised = revision_result.get("revised_hypotheses", hypotheses) or hypotheses

    elapsed = time.time() - start_time

    # --- Scoring ---
    top_names = [h.get("name", "") for h in revised[:5]]
    top1_name = top_names[0] if top_names else ""
    top3_names = top_names[:3]

    def is_match(exp: str, pred: str) -> bool:
        if not exp or not pred: return False
        e = exp.lower().strip()
        p = pred.lower().strip()
        if e == p: return True
        # Allow partial matches for parentheticals like "Stroke (CVA)" vs "Stroke"
        if e in p or p in e: return True
        # Explicit synonyms (bidirectional)
        synonyms = {
            "copd": ["chronic obstructive pulmonary disease"],
            "kidney stone": ["nephrolithiasis", "kidney stones", "nephrolithiasis (kidney stones)", "renal calculi"],
            "type 2 diabetes mellitus": ["diabetes mellitus", "type 2 diabetes", "diabetes mellitus type 2", "t2dm"],
            "atrial fibrillation (afib)": ["atrial fibrillation", "afib", "a-fib"],
            "sepsis": ["septicemia", "severe sepsis", "septic shock"],
            "depression": ["major depressive disorder", "major depression", "clinical depression", "depressive disorder"],
            "osteoarthritis": ["osteoarthritis of the knee", "degenerative joint disease", "oa"],
            "liver cirrhosis": ["cirrhosis", "hepatic cirrhosis", "alcoholic cirrhosis"],
            "hepatitis b": ["hepatitis", "viral hepatitis", "hbv infection"],
            "eczema": ["atopic dermatitis", "atopic eczema", "dermatitis"],
            "gastroesophageal reflux disease": ["gerd", "acid reflux disease", "reflux esophagitis"],
            "acute appendicitis": ["appendicitis"],
            "acute myocardial infarction": ["myocardial infarction", "heart attack", "mi", "stemi", "nstemi"],
            "bell's palsy": ["facial nerve palsy", "idiopathic facial palsy", "facial palsy"],
            "anaphylaxis": ["anaphylactic shock", "anaphylactic reaction", "severe allergic reaction"],
            "deep vein thrombosis": ["dvt", "venous thromboembolism"],
            "anxiety disorder": ["generalized anxiety disorder", "panic disorder", "gad", "panic attack"],
            "pulmonary embolism": ["pe", "pulmonary thromboembolism"],
            "acute pancreatitis": ["pancreatitis"],
            "rheumatoid arthritis": ["ra", "seropositive rheumatoid arthritis"],
            "influenza": ["flu", "seasonal influenza"],
            "pneumothorax": ["collapsed lung", "tension pneumothorax"],
            "tuberculosis": ["tb", "pulmonary tuberculosis", "pulmonary tb"],
            "celiac disease": ["coeliac disease", "celiac sprue", "gluten-sensitive enteropathy"],
            "meningitis": ["bacterial meningitis", "viral meningitis"],
        }
        # Check both directions
        for key, vals in synonyms.items():
            if (e == key and p in vals) or (p == key and e in vals):
                return True
            if e in vals and p == key:
                return True
            if p in vals and e == key:
                return True
        return False

    top1_match = is_match(ground_truth, top1_name)
    top3_match = any(is_match(ground_truth, n) for n in top3_names)
    top5_match = any(is_match(ground_truth, n) for n in top_names)

    top1_confidence = revised[0].get("confidence", 0) if revised else 0

    return {
        "case_id": case_id,
        "ground_truth": ground_truth,
        "severity": case["severity"],
        "system": case["system"],
        "difficulty": case["difficulty"],
        "predicted_top1": top1_name,
        "predicted_top3": top3_names,
        "predicted_top5": top_names,
        "top1_confidence": round(top1_confidence, 1),
        "top1_match": top1_match,
        "top3_match": top3_match,
        "top5_match": top5_match,
        "symptoms_extracted": len(normalized.get("primary_symptoms", [])) + len(normalized.get("secondary_symptoms", [])),
        "evidence_items": len(evidence_result.get("evidence_ledger", [])),
        "elapsed_seconds": round(elapsed, 2),
        "all_hypotheses": [
            {"name": h.get("name", ""), "confidence": round(h.get("confidence", 0), 1)}
            for h in revised
        ],
    }


def compute_metrics(results: List[Dict]) -> Dict[str, Any]:
    """Compute aggregate accuracy metrics from test results."""
    total = len(results)
    if total == 0:
        return {"error": "No test results"}

    top1_correct = sum(1 for r in results if r["top1_match"])
    top3_correct = sum(1 for r in results if r["top3_match"])
    top5_correct = sum(1 for r in results if r["top5_match"])

    # Per-severity breakdown
    severity_groups = {}
    for r in results:
        sev = r["severity"]
        if sev not in severity_groups:
            severity_groups[sev] = {"total": 0, "top1": 0, "top3": 0}
        severity_groups[sev]["total"] += 1
        if r["top1_match"]:
            severity_groups[sev]["top1"] += 1
        if r["top3_match"]:
            severity_groups[sev]["top3"] += 1

    severity_metrics = {}
    for sev, counts in severity_groups.items():
        severity_metrics[sev] = {
            "total_cases": counts["total"],
            "top1_accuracy": round(counts["top1"] / counts["total"] * 100, 1),
            "top3_accuracy": round(counts["top3"] / counts["total"] * 100, 1),
        }

    # Per-difficulty breakdown
    diff_groups = {}
    for r in results:
        d = r["difficulty"]
        if d not in diff_groups:
            diff_groups[d] = {"total": 0, "top1": 0, "top3": 0}
        diff_groups[d]["total"] += 1
        if r["top1_match"]:
            diff_groups[d]["top1"] += 1
        if r["top3_match"]:
            diff_groups[d]["top3"] += 1

    difficulty_metrics = {}
    for d, counts in diff_groups.items():
        difficulty_metrics[d] = {
            "total_cases": counts["total"],
            "top1_accuracy": round(counts["top1"] / counts["total"] * 100, 1),
            "top3_accuracy": round(counts["top3"] / counts["total"] * 100, 1),
        }

    # Failed cases
    failed_top1 = [
        {"case_id": r["case_id"], "ground_truth": r["ground_truth"], "predicted": r["predicted_top1"],
         "confidence": r["top1_confidence"], "severity": r["severity"]}
        for r in results if not r["top1_match"]
    ]
    failed_top3 = [
        {"case_id": r["case_id"], "ground_truth": r["ground_truth"], "predicted_top3": r["predicted_top3"],
         "severity": r["severity"]}
        for r in results if not r["top3_match"]
    ]

    avg_confidence_correct = 0
    avg_confidence_wrong = 0
    correct_confs = [r["top1_confidence"] for r in results if r["top1_match"]]
    wrong_confs = [r["top1_confidence"] for r in results if not r["top1_match"]]
    if correct_confs:
        avg_confidence_correct = round(sum(correct_confs) / len(correct_confs), 1)
    if wrong_confs:
        avg_confidence_wrong = round(sum(wrong_confs) / len(wrong_confs), 1)

    avg_time = round(sum(r["elapsed_seconds"] for r in results) / total, 2)

    return {
        "total_cases": total,
        "top1_accuracy": round(top1_correct / total * 100, 1),
        "top3_accuracy": round(top3_correct / total * 100, 1),
        "top5_accuracy": round(top5_correct / total * 100, 1),
        "top1_correct": top1_correct,
        "top3_correct": top3_correct,
        "top5_correct": top5_correct,
        "avg_confidence_when_correct": avg_confidence_correct,
        "avg_confidence_when_wrong": avg_confidence_wrong,
        "avg_time_per_case_seconds": avg_time,
        "by_severity": severity_metrics,
        "by_difficulty": difficulty_metrics,
        "failed_top1": failed_top1,
        "failed_top3": failed_top3,
    }


def print_report(metrics: Dict, results: List[Dict]):
    """Print a formatted console report."""
    print("\n" + "=" * 70)
    print("   CURABOT GOLDEN TEST FRAMEWORK — ACCURACY REPORT")
    print("=" * 70)

    print(f"\n  Total Test Cases:  {metrics['total_cases']}")
    print(f"  Top-1 Accuracy:    {metrics['top1_accuracy']}%  ({metrics['top1_correct']}/{metrics['total_cases']})")
    print(f"  Top-3 Accuracy:    {metrics['top3_accuracy']}%  ({metrics['top3_correct']}/{metrics['total_cases']})")
    print(f"  Top-5 Accuracy:    {metrics['top5_accuracy']}%  ({metrics['top5_correct']}/{metrics['total_cases']})")
    print(f"  Avg Confidence (correct): {metrics['avg_confidence_when_correct']}%")
    print(f"  Avg Confidence (wrong):   {metrics['avg_confidence_when_wrong']}%")
    print(f"  Avg Time per Case:        {metrics['avg_time_per_case_seconds']}s")

    print("\n  --- Accuracy by Severity ---")
    for sev, m in metrics["by_severity"].items():
        status = "PASS" if m["top3_accuracy"] >= 80 else "WARN" if m["top3_accuracy"] >= 60 else "FAIL"
        print(f"    {sev:12s}  Top1: {m['top1_accuracy']:5.1f}%  Top3: {m['top3_accuracy']:5.1f}%  ({m['total_cases']} cases)  [{status}]")

    print("\n  --- Accuracy by Difficulty ---")
    for diff, m in metrics["by_difficulty"].items():
        print(f"    {diff:15s}  Top1: {m['top1_accuracy']:5.1f}%  Top3: {m['top3_accuracy']:5.1f}%  ({m['total_cases']} cases)")

    if metrics["failed_top3"]:
        print(f"\n  --- Failed Cases (missed Top-3) ---")
        for f in metrics["failed_top3"]:
            print(f"    {f['case_id']}: Expected '{f['ground_truth']}', Got {f['predicted_top3']}")

    # Overall verdict
    overall = metrics["top3_accuracy"]
    if overall >= 90:
        verdict = "EXCELLENT"
    elif overall >= 80:
        verdict = "GOOD"
    elif overall >= 70:
        verdict = "ACCEPTABLE"
    elif overall >= 60:
        verdict = "NEEDS IMPROVEMENT"
    else:
        verdict = "FAILING"

    print(f"\n  OVERALL VERDICT: {verdict} (Top-3: {overall}%)")
    print("=" * 70 + "\n")


def save_json_report(metrics: Dict, results: List[Dict], output_path: str):
    """Save full report as JSON."""
    report = {
        "report_generated": datetime.utcnow().isoformat(),
        "framework_version": "1.0",
        "metrics": metrics,
        "individual_results": results,
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"  JSON report saved to: {output_path}")


def save_excel_report(metrics: Dict, results: List[Dict], output_path: str):
    """Save summary as Excel spreadsheet."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # --- Sheet 1: Summary ---
        ws = wb.active
        ws.title = "Summary"
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        pass_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        fail_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        ws.append(["CuraBot Golden Test Report"])
        ws["A1"].font = Font(size=14, bold=True)
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        ws.append([])
        ws.append(["Metric", "Value"])
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        summary_rows = [
            ("Total Cases", metrics["total_cases"]),
            ("Top-1 Accuracy", f"{metrics['top1_accuracy']}%"),
            ("Top-3 Accuracy", f"{metrics['top3_accuracy']}%"),
            ("Top-5 Accuracy", f"{metrics['top5_accuracy']}%"),
            ("Avg Confidence (Correct)", f"{metrics['avg_confidence_when_correct']}%"),
            ("Avg Confidence (Wrong)", f"{metrics['avg_confidence_when_wrong']}%"),
            ("Avg Time per Case", f"{metrics['avg_time_per_case_seconds']}s"),
        ]
        for row in summary_rows:
            ws.append(row)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # --- Sheet 2: Individual Results ---
        ws2 = wb.create_sheet("Results")
        headers = ["Case ID", "Ground Truth", "Severity", "Difficulty", "Predicted #1",
                    "Confidence", "Top-1 Match", "Top-3 Match", "Time (s)"]
        ws2.append(headers)
        for i, cell in enumerate(ws2[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for r in results:
            row = [
                r["case_id"], r["ground_truth"], r["severity"], r["difficulty"],
                r["predicted_top1"], r["top1_confidence"],
                "PASS" if r["top1_match"] else "FAIL",
                "PASS" if r["top3_match"] else "FAIL",
                r["elapsed_seconds"],
            ]
            ws2.append(row)
            row_num = ws2.max_row
            # Color the pass/fail cells
            for col in [7, 8]:
                cell = ws2.cell(row=row_num, column=col)
                cell.fill = pass_fill if cell.value == "PASS" else fail_fill
                cell.border = thin_border

        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws2.column_dimensions[col_letter].width = 18

        # --- Sheet 3: By Severity ---
        ws3 = wb.create_sheet("By Severity")
        ws3.append(["Severity", "Cases", "Top-1 Accuracy", "Top-3 Accuracy"])
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        for sev, m in metrics["by_severity"].items():
            ws3.append([sev, m["total_cases"], f"{m['top1_accuracy']}%", f"{m['top3_accuracy']}%"])

        wb.save(output_path)
        print(f"  Excel report saved to: {output_path}")

    except ImportError:
        print("  [SKIP] openpyxl not installed — skipping Excel report")


async def main():
    parser = argparse.ArgumentParser(description="CuraBot Golden Test Framework")
    parser.add_argument("--limit", type=int, default=None, help="Limit number of cases to run")
    parser.add_argument("--severity", type=str, default=None, help="Filter by severity class")
    parser.add_argument("--id", type=str, default=None, help="Run a single test by ID")
    parser.add_argument("--no-excel", action="store_true", help="Skip Excel report generation")
    args = parser.parse_args()

    cases = load_golden_cases(
        filter_severity=args.severity,
        filter_id=args.id,
        limit=args.limit,
    )
    print(f"\n  Loading {len(cases)} golden test cases...")

    results = []
    for i, case in enumerate(cases):
        status = f"  [{i+1}/{len(cases)}] Testing {case['id']}: {case['ground_truth']}..."
        print(status, end="", flush=True)
        try:
            result = await run_single_test(case)
            results.append(result)
            match_str = "TOP-1" if result["top1_match"] else ("TOP-3" if result["top3_match"] else "MISS")
            print(f" {match_str} (predicted: {result['predicted_top1']}, {result['top1_confidence']}%, {result['elapsed_seconds']}s)")
        except Exception as e:
            print(f" ERROR: {e}")
            results.append({
                "case_id": case["id"], "ground_truth": case["ground_truth"],
                "severity": case["severity"], "system": case["system"],
                "difficulty": case["difficulty"], "predicted_top1": "ERROR",
                "predicted_top3": [], "predicted_top5": [],
                "top1_confidence": 0, "top1_match": False,
                "top3_match": False, "top5_match": False,
                "symptoms_extracted": 0, "evidence_items": 0,
                "elapsed_seconds": 0, "all_hypotheses": [],
            })

    metrics = compute_metrics(results)
    print_report(metrics, results)

    # Save reports
    eval_dir = os.path.dirname(os.path.abspath(__file__))
    save_json_report(metrics, results, os.path.join(eval_dir, "golden_test_results.json"))
    if not args.no_excel:
        save_excel_report(metrics, results, os.path.join(eval_dir, "golden_test_report.xlsx"))


if __name__ == "__main__":
    asyncio.run(main())
