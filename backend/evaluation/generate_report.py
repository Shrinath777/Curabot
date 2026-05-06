"""Generate Excel report from the latest baseline results."""
import json, os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

eval_dir = os.path.dirname(os.path.abspath(__file__))
data = json.load(open(os.path.join(eval_dir, "golden_test_baseline.json"), "r"))
results = data["results"]
metrics = data["metrics"]

wb = openpyxl.Workbook()
hdr_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
pass_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
fail_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

# Sheet 1: Summary
ws = wb.active
ws.title = "Summary"
ws.append(["CuraBot Golden Test — Baseline Report"])
ws["A1"].font = Font(size=14, bold=True)
ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
ws.append([])
ws.append(["Metric", "Value"])
for c in ws[4]: c.fill, c.font, c.border = hdr_fill, hdr_font, thin
for row in [
    ("Total Cases", metrics["total"]),
    ("Top-1 Accuracy", f"{metrics['top1_pct']}%"),
    ("Top-3 Accuracy", f"{metrics['top3_pct']}%"),
    ("Top-5 Accuracy", f"{metrics['top5_pct']}%"),
    ("Mode", "Offline Baseline (Rule-Based Only)"),
]:
    ws.append(row)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 20

# Sheet 2: Results
ws2 = wb.create_sheet("Results")
headers = ["Case ID", "Ground Truth", "Severity", "Difficulty", "Predicted #1", "Confidence", "Top-1", "Top-3", "Symptoms", "Time(s)"]
ws2.append(headers)
for c in ws2[1]: c.fill, c.font, c.border = hdr_fill, hdr_font, thin
for r in results:
    ws2.append([
        r["case_id"], r["ground_truth"], r["severity"], r["difficulty"],
        r["predicted_top1"], r["top1_confidence"],
        "PASS" if r["top1_match"] else "FAIL",
        "PASS" if r["top3_match"] else "FAIL",
        r["symptoms_found"], r["elapsed_seconds"]
    ])
    row_num = ws2.max_row
    for col in [7, 8]:
        cell = ws2.cell(row=row_num, column=col)
        cell.fill = pass_fill if cell.value == "PASS" else fail_fill
        cell.border = thin
for letter in "ABCDEFGHIJ":
    ws2.column_dimensions[letter].width = 18

out = os.path.join(eval_dir, "golden_test_report.xlsx")
wb.save(out)
print(f"Excel report saved: {out}")
