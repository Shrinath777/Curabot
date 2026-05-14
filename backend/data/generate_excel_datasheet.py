import os
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

def generate_excel():
    data_dir = os.path.dirname(__file__)
    json_path = os.path.join(data_dir, "diseases.json")
    excel_path = os.path.join(data_dir, "Disease_Datasheet_Gov_Formatted.xlsx")
    
    if not os.path.exists(json_path):
        print(f"Error: Could not find {json_path}")
        return
        
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            diseases = json.load(f)
            
        # Process data
        processed_data = []
        for d in diseases:
            # Handle symptoms properly whether it's a list or dict
            symptoms = d.get("symptoms", [])
            if isinstance(symptoms, dict):
                symptoms_str = ", ".join(symptoms.get("primary", []) + symptoms.get("secondary", []))
            else:
                symptoms_str = ", ".join(symptoms)

            row = {
                "Disease Name": d.get("name", ""),
                "ICD-10 Code": d.get("icd10", ""),
                "System": d.get("system", ""),
                "Acuity": d.get("acuity", ""),
                "Prevalence": d.get("prevalence", ""),
                "Severity Class": d.get("severity_class", ""),
                "Description": d.get("description", ""),
                "Symptoms": symptoms_str,
                "Vital Sign Patterns": str(d.get("vital_sign_patterns", {})),
                "Lab Tests": ", ".join(d.get("lab_tests", [])),
                "Risk Factors": ", ".join(d.get("risk_factors", [])),
                "Red Flags": ", ".join(d.get("red_flags", [])),
                "Key Diagnostic Questions": " | ".join(d.get("key_diagnostic_questions", [])),
                "Differentiating Features": " | ".join(d.get("differentiating_features", [])),
                "References & Citations": ", ".join(d.get("references", [])),
                "EOI Reference": d.get("eoi_reference", "")
            }
            processed_data.append(row)
            
        df = pd.DataFrame(processed_data)
        df.to_excel(excel_path, index=False)
        
        # Apply formatting with openpyxl
        wb = load_workbook(excel_path)
        ws = wb.active
        
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
                             
        alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Format headers
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
        # Format rows and columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = alignment
                
        # Set column widths
        for col in ws.columns:
            column_letter = col[0].column_letter
            ws.column_dimensions[column_letter].width = 30 # standard width for content
            
        ws.column_dimensions['A'].width = 25 # Disease Name
        ws.column_dimensions['B'].width = 12 # ICD-10 Code
        ws.column_dimensions['G'].width = 45 # Description
        ws.column_dimensions['P'].width = 50 # EOI Reference
        
        wb.save(excel_path)
        
        print(f"Successfully generated and formatted {excel_path} with {len(processed_data)} diseases.")
        
        # Delete old clumsy files if they exist
        old_paths = [
            os.path.join(data_dir, "disease_datasheet_Gov.xlsx"),
            os.path.join(data_dir, "diseases_datasheet_GOV.xlsx"),
            os.path.join(data_dir, "diseases_datasheet.xlsx")
        ]
        for p in old_paths:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except:
                    pass
            
    except Exception as e:
        print(f"Error processing diseases.json: {e}")

if __name__ == "__main__":
    generate_excel()
